from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
DOCS_DIR = ROOT / "docs"
EXCEL_DIR = ROOT / "excel"
OUTPUT_PATH = EXCEL_DIR / "Azure-HA-DR-Dashboard.xlsx"

TAB_SPECS = [
    ("Executive Summary", DOCS_DIR / "executive-summary.md"),
    ("Interactive Dashboard", None),
    ("Service Catalog", DATA_DIR / "service_catalog_template.csv"),
    ("RTO-RPO Decision Matrix", DATA_DIR / "rto_rpo_decision_matrix.csv"),
    ("HA-DR Patterns", DATA_DIR / "research_knowledge_base.csv"),
    ("Architecture References", DATA_DIR / "architecture_references.csv"),
    ("MS Learn Links", DATA_DIR / "source_register.csv"),
    ("Glossary", DOCS_DIR / "glossary.md"),
    ("Assumptions & Caveats", DOCS_DIR / "assumptions-and-caveats.md"),
    ("Source Register", DATA_DIR / "source_register.csv"),
]

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
INFO_FILL = PatternFill(fill_type="solid", fgColor="EAF4FE")
TITLE_FONT = Font(bold=True)


def read_csv_rows(path: Path) -> list[list[str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.reader(handle))


def markdown_to_lines(path: Path) -> list[str]:
    with path.open("r", encoding="utf-8") as handle:
        return [line.rstrip("\n") for line in handle]


def autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        values = [str(cell.value) for cell in column_cells if cell.value is not None]
        max_length = max((len(value) for value in values), default=0)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 60)


def write_table(worksheet, rows: Iterable[Iterable[str]]) -> None:
    for row_index, row in enumerate(rows, start=1):
        for column_index, value in enumerate(row, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            if row_index == 1:
                cell.font = TITLE_FONT
                cell.fill = HEADER_FILL
    worksheet.freeze_panes = "A2"
    autosize_columns(worksheet)


def write_markdown_sheet(worksheet, lines: list[str]) -> None:
    for index, line in enumerate(lines, start=1):
        cell = worksheet.cell(row=index, column=1, value=line)
        if line.startswith("#"):
            cell.font = TITLE_FONT
            cell.fill = INFO_FILL
    worksheet.column_dimensions["A"].width = 120


def build_dashboard_sheet(worksheet) -> None:
    worksheet["A1"] = "Azure HA/DR Interactive Dashboard"
    worksheet["A1"].font = Font(bold=True, size=14)
    worksheet["A1"].fill = INFO_FILL

    prompts = [
        ("A3", "Select Azure Service"),
        ("A4", "Select Workload Category"),
        ("A5", "Select Criticality"),
        ("A6", "Desired RTO"),
        ("A7", "Desired RPO"),
        ("A8", "Region / Zone Preference"),
        ("A9", "Topology Preference"),
        ("A10", "Cost Sensitivity"),
        ("A11", "Simplicity vs Resilience"),
        ("D3", "Recommended HA Approach"),
        ("D4", "Recommended DR Approach"),
        ("D5", "Recommended Architecture Pattern"),
        ("D6", "Resiliency Classification"),
        ("D7", "Technical Reasoning"),
        ("D8", "Architecture References"),
        ("D9", "Best Practice Links"),
        ("D10", "Limitations and Risks"),
        ("D11", "Operational Guidance"),
        ("D12", "Leadership Summary"),
    ]

    for cell_ref, label in prompts:
        worksheet[cell_ref] = label
        worksheet[cell_ref].font = TITLE_FONT

    note = (
        "This starter workbook includes the dashboard layout only. "
        "Add named ranges, dropdown validation, and lookup formulas after the datasets are populated."
    )
    worksheet["A13"] = note
    worksheet.merge_cells("A13:F14")
    worksheet["A13"].fill = INFO_FILL

    worksheet.column_dimensions["A"].width = 28
    worksheet.column_dimensions["B"].width = 28
    worksheet.column_dimensions["D"].width = 30
    worksheet.column_dimensions["E"].width = 42
    worksheet.column_dimensions["F"].width = 42


def main() -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for sheet_name, source_path in TAB_SPECS:
        worksheet = workbook.create_sheet(title=sheet_name)
        if sheet_name == "Interactive Dashboard":
            build_dashboard_sheet(worksheet)
            continue

        if source_path is None:
            continue

        if source_path.suffix.lower() == ".csv":
            write_table(worksheet, read_csv_rows(source_path))
        else:
            write_markdown_sheet(worksheet, markdown_to_lines(source_path))

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)
    print(f"Workbook created: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
